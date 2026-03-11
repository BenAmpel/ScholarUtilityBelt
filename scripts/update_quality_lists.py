#!/usr/bin/env python3
"""
Fetch and build venue quality lists for local use in the extension.

Design constraints:
- Zero interaction with Google Scholar.
- Keep data local (written into src/data/*).

Notes:
- The UTD journals page currently fails TLS verification in this environment; we fetch it with
  an unverified SSL context (equivalent to curl --insecure). If you prefer, download manually
  and point the parser at a local file.
"""

from __future__ import annotations

import csv
import io
import json
import re
import ssl
import sys
import time
import urllib.request
from pathlib import Path
from typing import Iterable

import openpyxl
import requests
from bs4 import BeautifulSoup

ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = ROOT / "src" / "data"

FT50_URL = "https://www.ft.com/content/3405a512-5cbb-11e1-8f1f-00144feabdc0"
UTD24_URL = "https://jsom.utdallas.edu/the-utd-top-100-business-school-research-rankings/list-of-journals"
ABDC_2022_URL = "https://abdc.edu.au/wp-content/uploads/2023/05/ABDC-JQL-2022-v3-100523.xlsx"
ABS_2024_URL = "https://journalranking.org/"
CORE_PORTAL_BASE = "https://portal.core.edu.au/conf-ranks/"
CORE_SOURCE = "ICORE2026"


def die(msg: str) -> None:
    print(msg, file=sys.stderr)
    raise SystemExit(2)


def write_text(path: Path, text: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(text, encoding="utf-8")


def uniq_preserve(xs: Iterable[str]) -> list[str]:
    seen: set[str] = set()
    out: list[str] = []
    for x in xs:
        x = " ".join(str(x).split()).strip()
        if not x:
            continue
        k = x.casefold()
        if k in seen:
            continue
        seen.add(k)
        out.append(x)
    return out


def fetch_text(url: str, *, verify_tls: bool = True, timeout_s: int = 40, headers: dict | None = None) -> str:
    h = dict(headers) if headers else {}
    if "User-Agent" not in h:
        h.setdefault("User-Agent", "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36")
    if verify_tls:
        r = requests.get(url, timeout=timeout_s, headers=h)
        r.raise_for_status()
        return r.text

    # Local opt-out for problem TLS chains.
    ctx = ssl.create_default_context()
    ctx.check_hostname = False
    ctx.verify_mode = ssl.CERT_NONE
    req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
    with urllib.request.urlopen(req, context=ctx, timeout=timeout_s) as resp:
        data = resp.read()
    return data.decode("utf-8", errors="ignore")


def fetch_bytes(url: str, *, timeout_s: int = 60) -> bytes:
    r = requests.get(url, timeout=timeout_s)
    r.raise_for_status()
    return r.content


def build_ft50() -> list[str]:
    # Avoid parsing the entire FT HTML with BeautifulSoup; it's large.
    html = fetch_text(FT50_URL, verify_tls=True)

    for m in re.finditer(
        r'<script[^>]+type="application/ld\+json"[^>]*>\s*(\{.*?\})\s*</script>',
        html,
        flags=re.I | re.S,
    ):
        blob = m.group(1)
        try:
            data = json.loads(blob)
        except Exception:
            continue
        if not isinstance(data, dict):
            continue
        if data.get("@type") != "NewsArticle":
            continue

        body = data.get("articleBody")
        if not isinstance(body, str) or "The list below details the 50 journals" not in body:
            continue

        journals: list[str] = []
        for line in body.splitlines():
            lm = re.match(r"^\s*(\d+)\.\s*(.+?)\s*\*?\s*$", line)
            if lm:
                journals.append(lm.group(2).strip())
        journals = uniq_preserve(journals)
        if len(journals) != 50:
            die(f"FT50 parse error: expected 50, got {len(journals)}")
        return journals

    die("FT50 parse error: could not find JSON-LD NewsArticle articleBody")


def build_utd24() -> list[str]:
    html = fetch_text(UTD24_URL, verify_tls=False)
    soup = BeautifulSoup(html, "html.parser")
    divs = soup.select("div.journal_list")
    journals = [d.get("title") or d.get_text(" ", strip=True) for d in divs]
    journals = uniq_preserve(journals)
    if len(journals) != 24:
        die(f"UTD24 parse error: expected 24, got {len(journals)}")
    return journals


def build_abdc_2022_csv() -> str:
    content = fetch_bytes(ABDC_2022_URL)
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    if "2022 JQL" not in wb.sheetnames:
        die(f"ABDC parse error: expected sheet '2022 JQL' in {wb.sheetnames}")
    ws = wb["2022 JQL"]

    header_row = None
    col_title = None
    col_rank = None
    for r in range(1, 60):
        vals = [ws.cell(r, c).value for c in range(1, 60)]
        lower = [str(v).strip().lower() for v in vals if v is not None]
        if "journal title" in lower and "2022 rating" in lower:
            header_row = r
            # Map column indices.
            for c in range(1, 60):
                v = ws.cell(r, c).value
                if not isinstance(v, str):
                    continue
                if v.strip().lower() == "journal title":
                    col_title = c
                elif v.strip().lower() == "2022 rating":
                    col_rank = c
            break

    if not header_row or not col_title or not col_rank:
        die("ABDC parse error: could not locate 'Journal Title' and '2022 rating' columns")

    out = io.StringIO()
    w = csv.writer(out, lineterminator="\n")
    w.writerow(["Venue", "Rank"])

    seen = set()
    n = 0
    min_col = min(col_title, col_rank)
    max_col = max(col_title, col_rank)
    t_off = col_title - min_col
    r_off = col_rank - min_col

    # iter_rows is substantially faster than addressing cells by coordinate in a loop.
    for row in ws.iter_rows(
        min_row=header_row + 1,
        min_col=min_col,
        max_col=max_col,
        values_only=True,
    ):
        title = row[t_off]
        rank = row[r_off]
        if not title or not rank:
            continue
        title = " ".join(str(title).split()).strip()
        rank = " ".join(str(rank).split()).strip()
        if not title or not rank:
            continue
        k = title.casefold()
        if k in seen:
            continue
        seen.add(k)
        w.writerow([title, rank])
        n += 1

    if n < 500:
        die(f"ABDC parse error: suspiciously low row count: {n}")

    return out.getvalue()


def build_abs_2024_csv() -> str:
    """Fetch ABS Journal Ranking 2024 from journalranking.org and output Venue,Rank CSV."""
    html = fetch_text(ABS_2024_URL, verify_tls=True)
    soup = BeautifulSoup(html, "html.parser")
    table = soup.find("table")
    if not table:
        die("ABS 2024 parse error: no table found")
    rows = table.find_all("tr")
    if len(rows) < 2:
        die("ABS 2024 parse error: table has no data rows")
    # Header: ISSN | FIELD | TITLE | PUBLISHER | AJG2024 | AJG2021 | AJG2018
    header_cells = [th.get_text(" ", strip=True).upper() for th in rows[0].find_all(["th", "td"])]
    col_title = None
    col_ajg = None
    for i, h in enumerate(header_cells):
        if "TITLE" in h and "PUBLISHER" not in h:
            col_title = i
        if "AJG2024" in h or "AJG 2024" in h:
            col_ajg = i
    if col_title is None or col_ajg is None:
        die("ABS 2024 parse error: could not locate TITLE and AJG2024 columns")
    out = io.StringIO()
    w = csv.writer(out, lineterminator="\n")
    w.writerow(["Venue", "Rank"])
    seen = set()
    n = 0
    VALID_ABS = {"1", "2", "3", "4", "4*"}
    for tr in rows[1:]:
        cells = [td.get_text(" ", strip=True) for td in tr.find_all(["td", "th"])]
        if len(cells) <= max(col_title, col_ajg):
            continue
        title = " ".join(str(cells[col_title]).split()).strip()
        rank = " ".join(str(cells[col_ajg]).split()).strip().replace(" ", "")
        if not title or not rank:
            continue
        rank_lower = rank.lower()
        if rank_lower not in {"1", "2", "3", "4", "4*"}:
            continue
        k = title.casefold()
        if k in seen:
            continue
        seen.add(k)
        w.writerow([title, rank])
        n += 1
    if n < 200:
        die(f"ABS 2024 parse error: suspiciously low row count: {n}")
    return out.getvalue()


def build_core_csv() -> tuple[str, dict]:
    # CORE portal supports paging with a 'page' param. For ICORE2026 it appears to be 20 pages.
    # We'll discover the last page by looking at the pagination jumpPage('N') links.
    params = {"search": "", "by": "all", "source": CORE_SOURCE, "page": "1", "sort": "atitle"}
    r = requests.get(CORE_PORTAL_BASE, params=params, timeout=40)
    r.raise_for_status()

    html = r.text
    pages = [int(x) for x in re.findall(r"jumpPage\('?(\d+)'?\)", html)]
    last_page = max(pages) if pages else 1

    rows_out: list[tuple[str, str]] = []

    def parse_page(page_html: str) -> list[tuple[str, str, str]]:
        soup = BeautifulSoup(page_html, "html.parser")
        rows = soup.select("table tr")
        if not rows:
            return []
        data = []
        for tr in rows[1:]:
            tds = [td.get_text(" ", strip=True) for td in tr.find_all("td")]
            if len(tds) < 4:
                continue
            title, acronym, source, rank = tds[0], tds[1], tds[2], tds[3]
            data.append((title, acronym, rank))
        return data

    for page in range(1, last_page + 1):
        print(f"[core] fetch page {page}/{last_page}", flush=True)
        params["page"] = str(page)
        rp = requests.get(CORE_PORTAL_BASE, params=params, timeout=40)
        rp.raise_for_status()
        for title, acronym, rank in parse_page(rp.text):
            if not title or not rank:
                continue
            # Keep only meaningful ranks.
            if rank.strip().lower() in {"unranked", "other", "n/a"}:
                continue
            name = title.strip()
            syns = [name]
            if acronym and acronym.strip() and acronym.strip().lower() not in {"none", "-"}:
                syns.append(acronym.strip())
            venue = "|".join(uniq_preserve(syns))
            rows_out.append((venue, rank.strip()))

        # Be polite.
        time.sleep(0.2)

    # Emit in insertion order but unique.
    seen: set[str] = set()
    lines: list[list[str]] = [["Venue", "Rank"]]
    for venue, rank in rows_out:
        k = venue.casefold()
        if k in seen:
            continue
        seen.add(k)
        lines.append([venue, rank])

    out2 = io.StringIO()
    w2 = csv.writer(out2, lineterminator="\n")
    for row in lines:
        w2.writerow(row)

    meta = {"source": CORE_SOURCE, "pages": last_page, "rows": len(lines) - 1}
    return out2.getvalue(), meta


def main() -> int:
    DATA_DIR.mkdir(parents=True, exist_ok=True)

    meta: dict = {
        "fetchedAt": time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime()),
        "sources": {},
    }

    print("[ft50] fetch+parse", flush=True)
    ft50 = build_ft50()
    write_text(DATA_DIR / "ft50.txt", "\n".join(ft50) + "\n")
    meta["sources"]["ft50"] = {"url": FT50_URL, "count": len(ft50)}

    print("[utd24] fetch+parse", flush=True)
    utd24 = build_utd24()
    write_text(DATA_DIR / "utd24.txt", "\n".join(utd24) + "\n")
    meta["sources"]["utd24"] = {"url": UTD24_URL, "count": len(utd24), "tls_verify": False}

    print("[abdc] fetch+parse", flush=True)
    abdc_csv = build_abdc_2022_csv()
    write_text(DATA_DIR / "abdc2022.csv", abdc_csv)
    meta["sources"]["abdc2022"] = {"url": ABDC_2022_URL, "rows": abdc_csv.count("\n") - 1}

    print("[abs2024] fetch+parse", flush=True)
    abs_csv = build_abs_2024_csv()
    write_text(DATA_DIR / "abs2024.csv", abs_csv)
    meta["sources"]["abs2024"] = {"url": ABS_2024_URL, "rows": abs_csv.count("\n") - 1}

    print("[core] fetch+parse", flush=True)
    core_csv, core_meta = build_core_csv()
    write_text(DATA_DIR / "core_icore2026.csv", core_csv)
    meta["sources"]["core_icore2026"] = {"url": CORE_PORTAL_BASE, **core_meta}

    write_text(DATA_DIR / "quality_sources.json", json.dumps(meta, indent=2, sort_keys=True) + "\n")

    print("Wrote:")
    for p in ["ft50.txt", "utd24.txt", "abdc2022.csv", "abs2024.csv", "core_icore2026.csv", "quality_sources.json"]:
        print(" -", (DATA_DIR / p).relative_to(ROOT))

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
